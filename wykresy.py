import PySimpleGUI as sg
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import subprocess
import sys
import schedule
import time
import numpy as np
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import os
import shutil
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import win32com.client as win32
import pandas as pd
from datetime import datetime, timedelta
import datetime as dt
import time
import numpy as np
import xlrd
import urllib.request
import requests
import sys
import getpass
import json
import lxml
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import holidays
import holidays.countries

cale=[]
kwartaly=[]
msc=[]
print(cale)
df = pd.read_excel(r'abc.xlsx')


df = df.replace('-',np.nan)   #zamienia  "-" na Nan w komórkach gdzie nie ma ceny
df = df.astype({'DKR':float})  #zamienia kolumne DKR na floaty (dane były jako string)
df['kontrakt short'] = df['Kontrakt'].str.split("_").str[-1]       #Skraca nazwe kontraktu do uniwersalnego (dla base i peak) żeby je sparować
df['Data']=pd.to_datetime(df['Data'], format='%d-%m-%Y')
df['wolumen'] = [float(str(val).replace(u'\xa0','').replace(',','.')) for val in df['wolumen'].values]   #wyrzucenie dziwnych znaków z wolumenu i zamiana na float
df3 = df[['Data','DKR','typ','wolumen','kontrakt short']]  #stworzenie skróconego df bez zbędnych kolumn
df_base = df3[df3['typ'] == 'BASE']     #stworzenie df dla base
df_peak = df3[df3['typ'] == 'PEAK']
df_wsp = pd.merge(df_base,df_peak, on=['Data','kontrakt short'])  #połączenie df_base i df_peak dzieki temu można dodać kolumne ratio
df_wsp['ratio']=df_wsp['DKR_y']/df_wsp['DKR_x']  #kolumna z ratio

# Pętla do uzupełniania listy produktów, które znajdują sie w pliku zródłowym
for produkt in df['Kontrakt']:
    if "_Y-" in produkt:
        if produkt not in cale:
            cale.append(produkt)
    elif "_Q-" in produkt:
        if produkt not in kwartaly:
            kwartaly.append(produkt)
    elif "_M-" in produkt:
        if produkt not in msc:
            msc.append(produkt)
    else:
        continue

msc.sort()
kwartaly.sort()
cale.sort()
print(cale)
print(kwartaly)
print(msc)
# Funkcja rysujaca wykres w matplotlib
def draw_chart(produkt):
    df2=df[df['Kontrakt']==produkt]
    data=df2['Data']
    cena=(df2['DKR'])
    wolumen=(df2['wolumen'])
    # Tworzenie figury i osi
    fig, ax1 = plt.subplots()
    ax1.bar(data, wolumen, color='gray', alpha=0.5)
    ax1.set_ylabel('Wolumen')
    ax2 = ax1.twinx()

    ax2.plot(data, cena, marker='o', linestyle='-', color='blue')
    ax2.set_title(produkt)
    ax2.set_xlabel('Data')
    ax2.set_ylabel('cena')
    #
    #myFmt = mdates.DateFormatter('%d-%m-%Y')
    fig.autofmt_xdate(rotation=35, ha='right')    #rotuje daty wyświetlane pod wykresem
    plt.show()


sg.theme("Black") #gotowe motywy z kolorystyka do podejrzenia w internecie

label=sg.Text("Wykres liniowy DKR, wraz z słupkami z wolumenem")

Y_label=sg.Text("Y")
Y_combo=sg.Combo(cale, font=('Arial Bold', 14),  expand_x=True, enable_events=True, key='Y_droplist')

Q_label=sg.Text("Q")
Q_combo=sg.Combo(kwartaly, font=('Arial Bold', 14),  expand_x=True, enable_events=True, key='Q_droplist')

M_label=sg.Text("M")
M_combo=sg.Combo(msc, font=('Arial Bold', 14),  expand_x=True, enable_events=True, key='M_droplist')

download_button=sg.Button(button_text="zaciągnij brakujące dane z TGE",key="download")


window=sg.Window("Wykresy",
                 layout=[[label],
                        [Y_label,Y_combo,Q_label,Q_combo,M_label,M_combo],
                        [download_button]],
                        size=(800,100),element_justification='c')


while True:
    event,values=window.read()
    print(f'events: {event}')
    print(f'values: {values}')
    match event:
        case sg.WIN_CLOSED:  # co się stanie po zamknięciu okna gui
            break
        case 'Y_droplist':
            draw_chart(values['Y_droplist'])
        case 'Q_droplist':
            draw_chart(values['Q_droplist'])
        case 'M_droplist':
            draw_chart(values['M_droplist'])
        case 'download':
            window.Hide()  # ukrywa pierwotne okno Gui(niezaktualizowane)
            # TUTAJ WKLEJAM CAŁY KOD Z PLIKU MAIN (żeby pyinstaller miał 1 plik do przekompilowania)
            # znajduje ostatnią uzupełnioną datę w pliku
            wb = load_workbook(filename="abc.xlsx")
            ws = wb["a"]
            ostatni_wiersz = ws.max_row
            ostatnia_data = ws.cell(row=ostatni_wiersz,
                                    column=1).value  # uchwycona ostatnia data, dla której są dane w pliku excel
            pl_holidays = holidays.Poland()

            # kod  na ostatni dzień roboczy
            dzisiaj = dt.date.today()
            delta1 = dt.timedelta(days=1)
            delta2 = dt.timedelta(days=2)
            ostatni_dzien = dzisiaj - delta1

            if ostatni_dzien.weekday() == 5:
                ostatni_dzien = ostatni_dzien - delta1
            elif ostatni_dzien.weekday() == 6:
                ostatni_dzien = ostatni_dzien - delta2

            ostatni_dzien_str = str(ostatni_dzien)

            for x in range(len(pl_holidays)):
                if ostatni_dzien_str in pl_holidays:
                    ostatni_dzien = ostatni_dzien - delta1
                    if ostatni_dzien.weekday() == 5:
                        ostatni_dzien = ostatni_dzien - delta1
                    elif ostatni_dzien.weekday() == 6:
                        ostatni_dzien = ostatni_dzien - delta2
                    ostatni_dzien_str = str(ostatni_dzien)

            weekdays = [5, 6]
            data_poczatkowa = dt.datetime.strptime(ostatnia_data,
                                                   "%d-%m-%Y") + delta1  # trzeba do ostatniej daty dodać 1 dzień
            start_day = data_poczatkowa
            end_day = ostatni_dzien

            daterange = pd.date_range(start_day, end_day)
            for date in daterange:
                if date.weekday() not in weekdays and date.strftime("%Y-%m-%d") not in pl_holidays:
                    dzien = date.strftime("%d-%m-%Y")

                    sciezkaWebDriver = r"C:\Users\psliwa\PycharmProjects\Pobieranie_danych_tge\chromedriver.exe"  # do ściezki doklejam chromedriver, który wczesniej instaluje ze strony (sprawdź wersje chrome i sciągnij odpowiedni chromedriver)
                    # https://chromedriver.chromium.org/downloads     link do sciągniecia chromedrivera
                    # ileDni=1
                    driver = webdriver.Chrome(executable_path=sciezkaWebDriver)
                    url = 'https://tge.pl/energia-elektryczna-otf?dateShow=' + dzien + '&dateAction=prev'
                    page = driver.get(url)
                    time.sleep(1)
                    # -----BASE ---
                    df = pd.read_html(driver.page_source, header=0, decimal=",", thousands='.')
                    df[0] = df[0].drop('Unnamed: 1', axis=1)  # df[0] dla base   df[1] dla peak
                    base = df[0]
                    base = base.iloc[:-1]  # usunięcie ostatniego wiersza z tabeli (podsumowania)
                    base.insert(0, 'data', dzien)  # dodanie daty w pierwszej kolumnie(0-zerowej)
                    base[
                        'typ'] = "BASE"  # dodajemy kolumne Typ: "BASE" dla produktów z tabeli base, PEAK dla produktów z tabeli PEAK
                    # -----PEAK-----
                    df[1] = df[1].drop('Unnamed: 1', axis=1)
                    peak = df[1]
                    peak = peak.iloc[:-1]
                    peak.insert(0, 'data', dzien)
                    peak['typ'] = 'PEAK'

                    wb = load_workbook(filename="abc.xlsx")
                    ws = wb["a"]
                    for x in dataframe_to_rows(base, index=False, header=False):
                        ws.append(x)  # append dodaje dane do już istniejących w pliku
                    wb.save("abc.xlsx")
                    for x in dataframe_to_rows(peak, index=False, header=False):
                        ws.append(x)
                    wb.save("abc.xlsx")

                    driver.close()
                    driver.quit()
                    # KONIEC PLIKU MAIN
            subprocess.Popen([sys.executable, "wykresy.py"])  # uruchamiam GUI jeszcze raz,żeby zaciągniete
window.close()
